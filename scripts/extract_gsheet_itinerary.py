import openpyxl
import json

wb = openpyxl.load_workbook('Sing - Malay.xlsx')
# Ưu tiên sheet có tên chứa 'Lịch trình chi tiết'
sheet = wb.active
for s in wb.sheetnames:
    if 'chi tiết' in s.lower():
        sheet = wb[s]
        break

# Tìm dòng header thực sự (dòng có nhiều giá trị)
header_row_idx = 1
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), 1):
    non_none = [v for v in row if v is not None]
    if len(non_none) >= 3 and 'giờ' in str(row[0]).lower():
        header_row_idx = i
        headers = [str(cell).strip() if cell else None for cell in row]
        break

rows = []
for row in sheet.iter_rows(min_row=header_row_idx+1, values_only=True):
    row_dict = dict(zip(headers, row))
    # Bỏ qua dòng trống
    if any(row_dict.values()):
        rows.append(row_dict)

# Chỉ lấy các cột chính
main_cols = ['Giờ', 'Địa điểm', 'Hướng dẫn di chuyển', 'Ghi chú']
main_cols_lower = [c.lower() for c in main_cols]

# Chuẩn hóa header mapping
header_map = {}
for h in headers:
    if h:
        h_l = h.lower().replace('\n', ' ').replace('\r', '').strip()
        for c in main_cols_lower:
            if c in h_l:
                header_map[c] = h

filtered = [
    {c: r.get(header_map[c.lower()], '') for c in main_cols if c.lower() in header_map}
    for r in rows
]

with open('parsed_itinerary_detail.json', 'w', encoding='utf-8') as f:
    json.dump(filtered, f, ensure_ascii=False, indent=2)

print(f"Extracted {len(filtered)} rows. Columns: {list(header_map.values())}")
