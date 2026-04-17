import openpyxl
import json

wb = openpyxl.load_workbook('scripts/Sing - Malay.xlsx')
sheet = wb.active

# Try to find the sheet with the most relevant name
for s in wb.sheetnames:
    if 'chi tiết' in s.lower() or 'itinerary' in s.lower():
        sheet = wb[s]
        break



# Use header at row 8, data from row 9
header_row_idx = 8

# Normalize header names
raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx))]
headers = []
for h in raw_headers:
    if h is None:
        headers.append(None)
    else:
        headers.append(str(h).replace('\n', ' ').replace('\r', '').strip().lower())

rows = []
for row in sheet.iter_rows(min_row=header_row_idx+1, values_only=True):
    row_dict = dict(zip(headers, row))
    rows.append(row_dict)

# Always include these columns if present (normalized)
must_have = ['giờ', 'địa điểm', 'hướng dẫn di chuyển', 'ghi chú']
relevant_cols = [h for h in headers if h and any(x in h for x in must_have)]

filtered = [
    {k: v for k, v in r.items() if k in relevant_cols}
    for r in rows
]

with open('parsed_itinerary_notes.json', 'w', encoding='utf-8') as f:
    json.dump(filtered, f, ensure_ascii=False, indent=2)

print(f"Extracted {len(filtered)} rows. Relevant columns: {relevant_cols}")
