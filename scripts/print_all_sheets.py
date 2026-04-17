import openpyxl

wb = openpyxl.load_workbook('scripts/Sing - Malay.xlsx')
for s in wb.sheetnames:
    print(f"Sheet: {s}")
    sheet = wb[s]
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        print(i, row)
    print("-"*40)
